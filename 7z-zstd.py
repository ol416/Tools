import os
import math
import subprocess
import psutil
from blake3 import blake3
from pathlib import Path
import concurrent.futures
import time
from openpyxl import Workbook
import re
from openpyxl.styles import Font
import random
import string

def get_targets(root: str = './') -> list:
    """
    遍历当前目录，返回所有待压缩的目标（文件或文件夹），
    排除当前脚本、7z_result目录以及 log.xlsx 文件
    """
    current_dir = Path(__file__).parent
    script_name = Path(__file__).name
    exclude = {script_name, "7z_result", "log.xlsx"}
    targets = []
    for item in current_dir.iterdir():
        if item.name in exclude:
            continue
        if item.is_file() or item.is_dir():
            targets.append(item.resolve())
    return targets

def read_file_in_chunks(file_path, block_size=1024):
    """以块的形式读取文件数据"""
    with open(file_path, 'rb') as f:
        while True:
            chunk = f.read(block_size)
            if not chunk:
                break
            yield chunk

def format_size(size: int) -> str:
    """
    将字节数格式化为人类可读的单位
    """
    units = ['B', 'KB', 'MB', 'GB', 'TB', 'PB']
    def recursive_format(n, level=0):
        if n >= 1024 and level < len(units) - 1:
            return recursive_format(n / 1024, level + 1)
        else:
            return n, units[level]
    value, unit = recursive_format(size)
    return f"{value:.3f} {unit}"

def get_memory_info() -> dict:
    """获取当前内存使用情况"""
    vm = psutil.virtual_memory()
    return {"m_usage": vm.used, "m_free": vm.free, "m_per": vm.percent}

def get_memory_usecache(use_percent=2/3) -> int:
    free_memory = get_memory_info()['m_free']
    return int(round(free_memory * use_percent) / 64) * 64

def compute_blake3(file_path: str, block_size: int = 1024) -> str:
    """
    使用 blake3 计算文件哈希值（流式读取）
    """
    hasher = blake3()
    for chunk in read_file_in_chunks(file_path, block_size):
        hasher.update(chunk)
    return hasher.hexdigest()

def compress_target(target: Path, password: str, result_dir: str = '7z_result') -> Path:
    """
    压缩目标（文件或文件夹），返回压缩后的文件路径。
    如果目标为文件夹，则在命令中添加 -r- 禁用递归（只压缩文件夹本身）。
    """
    base = target.name
    result_path = target.parent / result_dir
    result_path.mkdir(exist_ok=True)
    result_file = result_path / f"{base}.7z"
    
    # 如果目标是文件夹，则禁用递归（-r-），否则不添加该参数
    rec_switch = " -r- " if target.is_dir() else " "
    
    # 7za 命令字符串：
    # -p{pwd}       设置密码
    # -mhe=on       启用加密文件头
    # -m0=bcj       启用 BCJ 过滤器
    # -m1=zstd      启用 zstd 压缩算法（当前版本支持）
    # -mx=11        最高压缩级别
    # -t7z         指定格式为 7z
    # -ms=64M      固实模式块大小
    # -slp         启用大页模式
    _7z_cmd = (
        '7za a{rec} -p{pwd} -mhe=on -m0=bcj -m1=zstd -mx=11 -t7z -ms=64M -slp "{result}" "{target}"'
    ).format(
        rec=rec_switch,
        pwd=password,
        result=str(result_file),
        target=str(target)
    )
    
    subprocess.run(_7z_cmd, shell=True, check=True)
    return result_file

def split_file(file_path: Path, threshold_mb: int) -> list:
    """
    将大于阈值的文件分割为多个部分，每个部分大小为设定阈值（MB）。
    返回分卷后生成的文件名列表，并删除原始文件。
    """
    file_size = file_path.stat().st_size
    threshold_bytes = threshold_mb * 1024 * 1024
    num_parts = math.ceil(file_size / threshold_bytes)
    part_files = []
    with open(file_path, 'rb') as f:
        for i in range(num_parts):
            part_file_name = file_path.with_name(f"{file_path.stem}{file_path.suffix}.{i+1:03d}")
            with open(part_file_name, 'wb') as pf:
                bytes_written = 0
                # 每个分卷严格写入 threshold_bytes 大小，最后一卷可能不足
                while bytes_written < threshold_bytes:
                    read_size = min(1024*1024, threshold_bytes - bytes_written)
                    data = f.read(read_size)
                    if not data:
                        break
                    pf.write(data)
                    bytes_written += len(data)
            part_files.append(part_file_name.name)
    # 删除原始完整文件
    file_path.unlink()
    return part_files

def process_target(target: Path, password: str, threshold_mb: int) -> dict:
    """
    对单个目标进行压缩，然后计算压缩文件大小和 Blake3 哈希，
    将压缩文件重命名为哈希值，若压缩文件体积超过阈值则自动执行分卷，
    返回记录字典。
    """
    start_time = time.time()
    compressed_file = compress_target(target, password)
    file_size = compressed_file.stat().st_size
    cache_block = get_memory_usecache()
    block_size = math.ceil(file_size / 64) * 64 if file_size < cache_block else cache_block
    hash_val = compute_blake3(str(compressed_file), block_size=block_size)
    size_str = format_size(file_size)
    elapsed = time.time() - start_time
    record = {
        "文件名": target.name,
        "blake3": hash_val,
        "pwd": password,
        "大小": size_str,
        "备注": f"耗时 {elapsed:.2f} 秒"
    }
    # 新增处理文件夹的子文件列表
    if target.is_dir():
        sub_files_info = []
        for path in target.rglob('*'):
            if path.is_file():
                # 修改相对路径基准为父目录，包含目标文件夹名称
                rel_path = path.relative_to(target.parent)
                file_size = path.stat().st_size
                sub_files_info.append({
                    "文件路径": str(rel_path),
                    "大小": format_size(file_size),
                })
        record["子文件列表"] = sub_files_info
    else:
        record["子文件列表"] = None  # 非文件夹默认设为None

    # 重命名压缩文件为 hash 值
    new_path = compressed_file.with_name(f"{hash_val}.7z")
    compressed_file.rename(new_path)
    record["压缩后文件"] = new_path.name

    # 检查是否超过分卷阈值，若超过则执行分卷
    if new_path.stat().st_size > threshold_mb * 1024 * 1024:
        parts = split_file(new_path, threshold_mb)
        record["分卷信息"] = parts
    else:
        record["分卷信息"] = None

    return record

def write_log_xlsx(records: list, filename: str = "log.xlsx"):
    # 对记录按类型（文件夹优先）和文件名排序
    records.sort(key=lambda x: (0 if x.get("子文件列表") else 1, x["文件名"]))
    
    wb = Workbook()
    # 主工作表
    ws_main = wb.active
    ws_main.title = "主日志"
    ws_main.append(["文件名", "blake3", "pwd", "大小", "备注"])
    
    # 存储其他工作表
    sub_sheets = {}
    
    for rec in records:
        # 写入主日志
        row = [
            rec["文件名"],
            rec["blake3"],
            rec["pwd"],
            rec["大小"],
            rec["备注"]
        ]
        ws_main.append(row)
        
        # 处理子文件列表
        sub_files = rec.get("子文件列表", None)
        if sub_files:
            # 设置文件名列的超链接
            row_num = ws_main.max_row
            cell = ws_main.cell(row=row_num, column=1)
            # 替换非法字符生成安全的工作表名称
            raw_name = rec["文件名"][:60]
            sheet_name = re.sub(r'[\[\]\\\*\:\?\/\|\"]', '_', raw_name)
            cell.hyperlink = f"#{sheet_name}!A1"
            cell.font = Font(underline='single', color="0563C1")  # 蓝色下划线
            
            # 创建或获取子工作表
            if sheet_name in sub_sheets:
                ws = sub_sheets[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
                ws.append(["文件路径", "大小"])  
                sub_sheets[sheet_name] = ws
            for sub_file in sub_files:
                ws.append([sub_file["文件路径"], sub_file["大小"]])
    
    wb.save(filename)

def main():
    # 修改：解析输入，支持"密码 分卷阈值"格式，如果未提供阈值则默认为 1024MB
    user_input = input("请输入密码或输入 'r' 生成随机密码及可选的分卷阈值(单位MB)，例如: r 500，或 mypassword 500，不输入则默认为 'ol416' 和 1024MB: ").strip()
    tokens = user_input.split()
    if tokens:
        first_token = tokens[0].lower()
        if first_token == 'r':
            password = ''.join(random.choices(string.ascii_letters + string.digits, k=18))
            if len(tokens) > 1 and tokens[1].isdigit():
                threshold_mb = int(tokens[1])
            else:
                threshold_mb = 1024
        else:
            password = tokens[0]
            if len(tokens) > 1 and tokens[1].isdigit():
                threshold_mb = int(tokens[1])
            else:
                threshold_mb = 1024
    else:
        password = 'ol416'
        threshold_mb = 1024

    targets = get_targets()
    records = []
    
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_to_target = {executor.submit(process_target, t, password, threshold_mb): t for t in targets}
        for future in concurrent.futures.as_completed(future_to_target):
            t = future_to_target[future]
            try:
                record = future.result()
                records.append(record)
                print(f"处理完成: {record['文件名']} -> {record.get('压缩后文件', '')}, "
                      f"Hash: {record['blake3']}, 大小: {record['大小']}, {record['备注']}")
                if record.get("分卷信息"):
                    print(f"分卷后文件: {record['分卷信息']}")
            except Exception as e:
                print(f"处理 {t} 时出错: {e}")
    
    write_log_xlsx(records)
    print("日志已保存为 log.xlsx")
    
    print("\n所有处理记录：")
    for rec in records:
        print(f"{rec['文件名']}\t{rec['blake3']}\t{rec['pwd']}\t{rec['大小']}\t{rec['备注']}")
        if rec.get("分卷信息"):
            print(f"分卷文件: {rec['分卷信息']}")

if __name__ == '__main__':
    main()
